import sys
import re
from collections import Counter
from openpyxl import load_workbook

def extract_three_char_chinese(text):
    return re.findall(r'[\u4e00-\u9fff]{3}', text)

def extract_text_from_xlsx(file_path):
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"載入 Excel 檔案失敗: {e}")
        return []

    all_text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    all_text.append(cell.value)

    return all_text

def main(file_path):
    all_lines = extract_text_from_xlsx(file_path)
    names = []

    for line in all_lines:
        names.extend(extract_three_char_chinese(line))

    counter = Counter(names)

    # print("重複的三字中文名字如下：")
    found = False
    for name, count in counter.items():
        if count > 1:
            print(f"{name}（共出現 {count} 次）")
            found = True

    # if not found:
    #     print("未找到重複的三字中文名字。")

if __name__ == "__main__":
    if len(sys.argv) != 2 or not sys.argv[1].endswith(".xlsx"):
        print("用法: python check_xlsx_names.py <檔案.xlsx>")
        sys.exit(1)

    main(sys.argv[1])
